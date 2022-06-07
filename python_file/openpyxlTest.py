from openpyxl import Workbook

wb = Workbook()
wb.create_sheet('0st_sheet', 0)
wb.create_sheet('1st_sheet', 1)
wb.create_sheet('2nd_sheet', 1)

ws = wb['0st_sheet']
ws['A1'] = 0

ws = wb['1st_sheet']
ws['A1'] = 1

ws = wb['2nd_sheet']
ws['A1'] = 2

wb.save(filename= "test.xlsx")
