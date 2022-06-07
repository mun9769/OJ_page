from selenium import webdriver
from openpyxl import Workbook
# driver = webdriver.Chrome('C://Users//New//Desktop//OJ_page//python_file//chromedriver.exe')
driver = webdriver.Chrome('C://Users//user//Desktop//OJ_page//python_file//chromedriver.exe')
url = 'https://sugang.inha.ac.kr/sugang/SU_51001/Lec_Time_Search.htm'
driver.get(url)

iframe = driver.find_element_by_id("ifmdtl")
driver.switch_to.frame(iframe)

wb = Workbook()

############# table name #################
test = driver.find_element_by_tag_name("thead")
title_row = test.find_element_by_css_selector("*")
title_col = title_row.find_elements_by_css_selector("*")
# for ele in title_col:
    # print(ele.get_attribute("innerText"))
##########################################

############# select #################
select_section = driver.find_element_by_id('ddlDept')
select_btn = driver.find_element_by_id("ibtnSearch1")

majors = select_section.find_elements_by_tag_name('option')
majors_len = len(majors)

for count in range(majors_len):
    driver.execute_script("document.querySelector('#ddlDept').children[{}].setAttribute('selected', 'selected')".format(count))


    select_btn = driver.find_element_by_id("ibtnSearch1")
    select_btn.click()

    select_section = driver.find_element_by_id('ddlDept')
    majors = select_section.find_elements_by_tag_name('option')
    sheet_name = majors[count].get_attribute('innerText')
    wb.create_sheet(str(count)) # 한글 지원안함. 

    ws = wb[str(count)]
    # print(type(sheet_name))
    # print(type(sheet_name.encode('utf-8').decode('utf-8')))

    # iframe = driver.find_element_by_tag_name('iframe')
    # driver.switch_to.frame(iframe)

    tbody = driver.find_element_by_tag_name('tbody')
    rows = tbody.find_elements_by_tag_name('tr')

    for i, row in enumerate(rows):
        cols = row.find_elements_by_tag_name('td')
        if i > 23: break
        for j, col in enumerate(cols):
            index_string = chr(65+ i) + str(j+1)
            ws[index_string] = col.get_attribute('innerText')
            # print(col.get_attribute('innerText'))
    
    a = 10
##########################################

wb.save(filename='hawing.xlsx')


# import requests

# req = requests.get('https://sugang.inha.ac.kr/sugang/SU_51001/Lec_Time_Search.htm')
# print(req.text)